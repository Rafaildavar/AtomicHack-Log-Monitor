"""Генератор отчетов - логика коллеги без изменений.

Содержит точную логику создания Excel отчетов из src/bot/services/analysis_history.py
"""

import logging
import os
import time
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Генератор Excel отчетов для результатов анализа.
    
    Логика написана коллегой и используется без изменений.
    """

    def __init__(self):
        """Инициализация генератора отчетов."""
        pass

    def create_excel_report(self, analysis_results: list, output_path: str = None) -> str:
        """Создает Excel отчет из результатов анализа в формате ValidationCases.xlsx.
        
        Точная логика коллеги из src/bot/services/analysis_history.py

        Args:
            analysis_results: Список результатов анализа файлов
            output_path: Путь для сохранения Excel файла (если None, генерируется уникальное имя)

        Returns:
            Путь к созданному файлу
        """
        try:
            # Генерируем уникальное имя файла с временной меткой, если путь не указан
            if output_path is None:
                timestamp = int(time.time())
                output_path = f"reports/file_analysis_{timestamp}.xlsx"
            
            # Создаем директорию если нужно
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Создаем DataFrame из результатов в формате ValidationCases
            if not analysis_results:
                # Создаем пустой отчет
                report_data = [{
                    'ID сценария': 0,
                    'ID аномалии': 0,
                    'ID проблемы': 0,
                    'Файл с проблемой': 'Нет данных',
                    '№ строки': 0,
                    'Строка из лога': 'Анализ не выполнен'
                }]
            else:
                report_data = []
                scenario_id_mapping = {}  # Для сопоставления имени сценария с ID
                current_scenario_id = 1
                
                for result in analysis_results:
                    if 'results' in result and result['results']:
                        # ML анализ с результатами - основной режим
                        for anomaly in result['results']:
                            # Получаем имя сценария
                            scenario_name = anomaly.get('Сценарий', 'Unknown')
                            
                            # Назначаем ID сценарию если его еще нет
                            if scenario_name not in scenario_id_mapping:
                                scenario_id_mapping[scenario_name] = current_scenario_id
                                current_scenario_id += 1
                            
                            scenario_id = scenario_id_mapping[scenario_name]
                            
                            # Формируем строку в точном формате ValidationCases.xlsx
                            report_data.append({
                                'ID сценария': scenario_id,
                                'ID аномалии': anomaly.get('ID аномалии', ''),
                                'ID проблемы': anomaly.get('ID проблемы', ''),
                                'Файл с проблемой': anomaly.get('Файл с проблемой', ''),
                                '№ строки': anomaly.get('№ строки', ''),
                                'Строка из лога': anomaly.get('Строка из лога', '')
                            })

            # Создаем DataFrame с точными колонками как в ValidationCases.xlsx
            df = pd.DataFrame(report_data)
            
            # Убеждаемся, что колонки в правильном порядке
            column_order = ['ID сценария', 'ID аномалии', 'ID проблемы', 
                          'Файл с проблемой', '№ строки', 'Строка из лога']
            df = df[column_order]
            
            # Сохраняем в Excel с форматированием
            # Сначала сохраняем DataFrame
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Загружаем workbook для форматирования
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Цвет заголовков (голубой как на скриншоте)
            header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            
            # Форматируем заголовки
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Выравнивание по центру для всех ячеек
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Настраиваем ширину колонок
            ws.column_dimensions['A'].width = 12  # ID сценария
            ws.column_dimensions['B'].width = 12  # ID аномалии
            ws.column_dimensions['C'].width = 12  # ID проблемы
            ws.column_dimensions['D'].width = 20  # Файл с проблемой
            ws.column_dimensions['E'].width = 10  # № строки
            ws.column_dimensions['F'].width = 80  # Строка из лога
            
            # Сохраняем отформатированный файл
            wb.save(output_path)
            wb.close()

            logger.info(f"Excel отчет создан: {output_path} ({len(df)} строк)")
            return output_path

        except Exception as e:
            logger.error(f"Ошибка создания Excel отчета: {e}", exc_info=True)
            return ""

